#功能：数据预处理，消除工频，去除坏道，全局平均重参考，数据保存为.mat和.fif格式
#作者：余江伟
#时间：2023/11/16; 2023/11/30:添加了基线校正和z-score标准化
#联系方式：15300593720

import mne
import numpy as np
import matplotlib.pyplot as plt
import json
import os
import yaml
from openpyxl import load_workbook
import hdf5storage as hdf
import pandas as pd
from scipy.stats import zscore
from scipy.signal import welch
from sklearn.preprocessing import StandardScaler


#reading yaml file, get parameters 读取yaml文件，获取参数
with open("D:\qq文件\交接代码\数据集\data\dataprocess.yaml", 'r',encoding='UTF-8') as f:
    data = yaml.load(f, Loader=yaml.FullLoader)

dataset = data.get('dataset')#data sources 数据来源
bad_channels_path = data.get('Badchannel_path')# bad channels file 记录坏道文件
high_freq = data.get('high_freq')#low pass filter frequcency低通滤波参数
low_freq = data.get('low_freq')#high pass filter frequency 高通滤波参数
onset_shift = data.get('onsetshift')#cropping onset shift time 裁剪onset时的偏移量
subject_num = data.get('sub_num')#subject number 数据subject数量
maxrun = data.get('maxrun')#max run numbers for subjects 每个subject最大试验次数
data_root = data.get('root')#data root path 数据根目录
bad_channel_start = data.get('bad_start')#certain data source's starting num of the bad channels file 某种数据坏道文件中记录的起点
saveroot = data.get('save_root')#processed data saving path 处理后数据的保存路径

#open bad channels file打开坏道文件
book = load_workbook(bad_channels_path)
sheet1 = book.get_sheet_by_name('Sheet1')

#choose 1 of 4 kind of data based on yaml parameter 根据参数选择某种数据来源
read_folder_path = " "
if dataset=="jh":
    read_folder_path = "/sub-jh10"
elif dataset=="ummc":
    read_folder_path = "/sub-ummc00"
elif dataset=="pt":
    read_folder_path = "/sub-pt"
elif dataset=="umf":
    read_folder_path = "/sub-umf00"

#loop for all subjects and runs 遍历该数据的所有受试者的所有实验
for i in range(1,subject_num+1):
    file_path = data_root+read_folder_path+"{}\ses-presurgery\ieeg".format(i)
    for j in range(1,maxrun+1):
        data_name = read_folder_path+"{}".format(i)+"_ses-presurgery_task-ictal_acq-ecog_run-0{}_ieeg".format(j)
        vhdr_data = file_path+data_name+".vhdr"
        json_data = file_path+data_name+".json"
        tsv_data = file_path+read_folder_path+"{}".format(i)+"_ses-presurgery_task-ictal_acq-ecog_run-0{}_events.tsv".format(j)
        #from .vhdr reading eeg file 从vhdr文件读取eeg文件
        if not(os.path.exists(vhdr_data)):
            continue
        raw = mne.io.read_raw_brainvision(vhdr_data, preload=True)
        #from .json reading powerline frequency and sampling frequceny 从json文件读取工频和采样率
        with open(json_data,'r',encoding='UTF-8') as f:
            json_info = json.load(f)
        PLF = json_info["PowerLineFrequency"]
        SF = round(json_info["SamplingFrequency"])
        #set default onset and offset 设置onset和offset的默认值
        onset = 10
        offset = int(raw.tmax)
        onsetsamp = 5000
        offsetsamp = 6000
        #for ummc data, read the annotations for onset and offset 对于ummc文件，直接读取标注获得onset和offset
        if dataset == "ummc":
            df = pd.read_csv(tsv_data, delimiter='\t')
            onset = df.loc[0]["onset"]#不算表头
            onsetsamp = df.loc[0]["sample"]
            offset = df.loc[1]["onset"]#不算表头
            offsetsamp = df.loc[1]["sample"]
        #for umf data, pt data and jh data, read .tsv file for onset and offset 对于umf,pt,jh数据，读取tsv文件获得onset和offset
        if dataset == "umf":
            df = pd.read_csv(tsv_data, delimiter='\t')
            for row in df.index:
                if df.loc[row]["trial_type"] == "eeg sz start":
                    onset = df.loc[row]["onset"]
                    onsetsamp = df.loc[row]["sample"]
                if df.loc[row]["trial_type"] == "eeg sz end":
                    offset = df.loc[row]["onset"]
                    offsetsamp = df.loc[row]["sample"]
                    break
        if dataset == "pt" or dataset == "jh":
            df = pd.read_csv(tsv_data, delimiter='\t')
            for row in df.index:
                if df.loc[row]["trial_type"] == "onset":
                    onset = df.loc[row]["onset"]
                    onsetsamp = df.loc[row]["sample"]
                if df.loc[row]["trial_type"] == "offset":
                    offset = df.loc[row]["onset"]
                    offsetsamp = df.loc[row]["sample"]
                    break
        raw.plot(scalings=4e-4,n_channels=60)
        plt.tight_layout()
        plt.show(block=True)#show raw data 展示原数据
        spectrum = raw.compute_psd()
        spectrum.plot()
        plt.tight_layout()
        plt.show(block=True)

        hl_data = raw.copy().filter(h_freq=high_freq, l_freq=low_freq)  # high-low pass filter 高-低通滤波
        # hl_data.plot(scalings=4e-4,n_channels=60)
        # plt.tight_layout()
        # plt.show(block=True)  # show data after high-low pass filter 展示高-低通滤波后的数据
        # spectrum_hl = hl_data.compute_psd()
        # spectrum_hl.plot()
        # plt.tight_layout()
        # plt.show(block=True)

        notch_data = hl_data.copy().notch_filter(freqs=PLF)  # removing powerline noise 去除工频噪音
        # notch_data.plot(scalings=4e-4,n_channels=60)
        # plt.tight_layout()
        # plt.show(block=True)  # show data after removing powerline noise 展示去除工频后的数据
        # spectrum_pl = notch_data.compute_psd()
        # spectrum_pl.plot()
        # plt.tight_layout()
        # plt.show(block=True)

        # drop bad channels based on the record 根据记录文件去除坏道
        drop_ch_data = notch_data.copy()
        if sheet1.cell(row=bad_channel_start + i, column=2).value != None:
            drop_chan = sheet1.cell(row=bad_channel_start + i, column=2).value.split(",")
            if dataset == "ummc":
                drop_chan.append('EVENT')
            drop_ch_data = drop_ch_data.drop_channels(drop_chan)
            # drop_ch_data.plot(scalings=4e-4,n_channels=60)
            # plt.tight_layout()
            # plt.show(block=True)  # show data after cropping and dropping channels 展示截取后的数据
            # spectrum_drop = drop_ch_data.compute_psd()
            # spectrum_drop.plot()
            # plt.tight_layout()
            # plt.show(block=True)

        # rereferencing with common average reference  通过全局平均参考进行重参考
        rereferenced_data, ref_data = mne.set_eeg_reference(drop_ch_data, copy=True)
        # rereferenced_data.plot(scalings=4e-4,n_channels=60)
        # plt.tight_layout()
        # plt.show(block=True)  # show the data after rereferencing 展示重参考后的数据
        # spectrum_ref = rereferenced_data.compute_psd()
        # spectrum_ref.plot()
        # plt.tight_layout()
        # plt.show(block=True)

        # in case there are less than 10 seconds shift before onset 避免有的run中onset前不足10秒
        # crop_data = rereferenced_data.copy().crop(tmin=max(onset + onset_shift, 0), tmax=offset)
        # onset_annotation = 10  # marking the onset to the processed data 为处理后的文件记录新的onset位置
        # if max(onset + onset_shift, 0) == 0:
        #     onset_annotation = onset
        # crop_data.plot(scalings=4e-4,n_channels=60)
        # plt.tight_layout()
        # plt.show(block=True)  # show data after cropping and dropping channels 展示截取后的数据
        # spectrum_crop = crop_data.compute_psd()
        # spectrum_crop.plot()
        # plt.tight_layout()
        # plt.show(block=True)

        # baseline correction 基线校正
        # baseline_start = onset_annotation - 5
        # baseline_end = onset_annotation - 1
        # if max(onset_annotation - 5, 0) == 0:
        #     baseline_start = 0
        baseline_start = -5
        baseline_end = -1
        tmin, tmax = onset_shift, offset - onset
        onset_annotation = -onset_shift  # marking the onset to the processed data 为处理后的文件记录新的onset位置
        if max(onset + onset_shift, 0) == 0:
            onset_annotation = onset
            tmin = -onset
        if max(onset_annotation - 5, 0) == 0:
            baseline_start = -onset_annotation
        baseline = (baseline_start, baseline_end)
        events = [[onsetsamp, 0, 1]]
        event_id = {'Event1': 1}

        # epochs = mne.make_fixed_length_epochs(crop_data, duration=int(crop_data.tmax),preload=True)#make one epoch 做成1个epoch
        epochs = mne.Epochs(rereferenced_data, events, event_id=event_id, tmin=tmin, tmax=tmax, baseline=baseline, preload=True)
        # epochs.apply_baseline(baseline)

        # Z-score standardization z-score标准化
        epoch_data = epochs.get_data()
        converting_data = np.squeeze(epoch_data)  # Reshape to (n_channels * n_times)
        # Calculate mean and standard deviation across channels(equals to in each row) 对每行数据（每个通道）做标准化
        epochs_data_standardized = zscore(converting_data,axis=1)

        # converting epoch data to raw data 将epoch 转为 raw
        info = epochs.info
        new_raw = mne.io.RawArray(epochs_data_standardized, info)
        new_raw.plot(scalings=1,n_channels=129)
        plt.tight_layout()
        plt.show(block=True)


        #record the new parameters 记录处理后数据的参数
        ch_num = new_raw.info['nchan']
        sample_num = new_raw.times.shape[0]
        duration = round(new_raw.tmax)


        #save .fif and .mat file 保存文件为fif文件和mat文件
        save_folder_path = saveroot+"/"+dataset+"/sub{}".format(i)
        if not os.path.exists(save_folder_path):
            os.mkdir(save_folder_path)
        save_path_fif = save_folder_path + "/sub0{}_".format(i)+"run0{}_data.fif".format(j)
        save_path_mat = save_folder_path + "/sub0{}_".format(i)+"run0{}_data.mat".format(j)
        rereferenced_data.save(save_path_fif,overwrite=True)
        hdf.savemat(file_name = save_path_mat,
                    mdict= {
                        "data":rereferenced_data.get_data(),#data array after processing 处理后的数据
                        "onset_annotation":onset_annotation,#onset annotation onset标注
                        'SamplingFrequency': SF,
                        'channel_num': ch_num,
                        'sample_num': sample_num,#采样个数
                        'duration': duration
                    })
